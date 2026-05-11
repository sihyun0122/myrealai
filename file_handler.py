"""
file_handler.py
- 이미지, PDF, Word, Excel, 텍스트 등 다양한 파일 처리
- Claude API로 전송할 수 있는 형태로 변환
"""

import base64
import io
import streamlit as st
from typing import Tuple, Optional

# PDF 처리
try:
    from PyPDF2 import PdfReader
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

# Word 처리
try:
    from docx import Document as DocxDocument
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

# Excel 처리
try:
    import openpyxl
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

# 이미지 처리
try:
    from PIL import Image
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False


# 지원 파일 형식
IMAGE_TYPES = ["image/jpeg", "image/png", "image/gif", "image/webp"]
DOCUMENT_TYPES = [
    "application/pdf",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "text/plain",
    "text/csv",
    "text/markdown",
    "application/json",
]
ALL_TYPES = IMAGE_TYPES + DOCUMENT_TYPES


def get_file_type_label(mime_type: str) -> str:
    """파일 타입 한글 이름"""
    labels = {
        "image/jpeg": "🖼️ JPEG 이미지",
        "image/png": "🖼️ PNG 이미지",
        "image/gif": "🖼️ GIF 이미지",
        "image/webp": "🖼️ WebP 이미지",
        "application/pdf": "📄 PDF 문서",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "📝 Word 문서",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "📊 Excel 문서",
        "text/plain": "📃 텍스트 파일",
        "text/csv": "📊 CSV 파일",
        "text/markdown": "📝 마크다운 파일",
        "application/json": "📋 JSON 파일",
    }
    return labels.get(mime_type, "📎 파일")


def is_image(mime_type: str) -> bool:
    """이미지 파일인지 확인"""
    return mime_type in IMAGE_TYPES


def extract_text_from_pdf(file_bytes: bytes) -> str:
    """PDF에서 텍스트 추출"""
    if not PDF_AVAILABLE:
        return "[PDF 라이브러리가 설치되지 않았습니다]"
    try:
        reader = PdfReader(io.BytesIO(file_bytes))
        texts = []
        for i, page in enumerate(reader.pages):
            text = page.extract_text()
            if text:
                texts.append(f"--- 페이지 {i+1} ---\n{text}")
        return "\n\n".join(texts) if texts else "[PDF에서 텍스트를 추출할 수 없습니다]"
    except Exception as e:
        return f"[PDF 읽기 오류: {e}]"


def extract_text_from_docx(file_bytes: bytes) -> str:
    """Word 문서에서 텍스트 추출"""
    if not DOCX_AVAILABLE:
        return "[Word 라이브러리가 설치되지 않았습니다]"
    try:
        doc = DocxDocument(io.BytesIO(file_bytes))
        texts = [para.text for para in doc.paragraphs if para.text.strip()]
        return "\n\n".join(texts) if texts else "[문서가 비어있습니다]"
    except Exception as e:
        return f"[Word 읽기 오류: {e}]"


def extract_text_from_xlsx(file_bytes: bytes) -> str:
    """Excel 파일에서 텍스트 추출"""
    if not XLSX_AVAILABLE:
        return "[Excel 라이브러리가 설치되지 않았습니다]"
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
        all_text = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_text.append(f"=== 시트: {sheet_name} ===")
            for row in ws.iter_rows(values_only=True):
                row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                if row_text.strip():
                    all_text.append(row_text)
        wb.close()
        return "\n".join(all_text) if all_text else "[Excel이 비어있습니다]"
    except Exception as e:
        return f"[Excel 읽기 오류: {e}]"


def process_uploaded_file(uploaded_file) -> Tuple[Optional[dict], str, str]:
    """
    업로드된 파일을 처리
    
    Returns:
        (claude_content, extracted_text, display_info)
        - claude_content: Claude API에 보낼 content 블록 (dict) 또는 None
        - extracted_text: 추출된 텍스트 (DB 저장용)
        - display_info: 화면에 표시할 정보 문자열
    """
    file_bytes = uploaded_file.read()
    mime_type = uploaded_file.type or "application/octet-stream"
    file_name = uploaded_file.name
    file_size = len(file_bytes)

    size_str = format_file_size(file_size)
    display_info = f"{get_file_type_label(mime_type)} | {file_name} | {size_str}"

    # ---- 이미지 처리 ----
    if is_image(mime_type):
        b64_data = base64.standard_b64encode(file_bytes).decode("utf-8")
        claude_content = {
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": mime_type,
                "data": b64_data,
            },
        }
        return claude_content, f"[이미지: {file_name}]", display_info

    # ---- PDF ----
    if mime_type == "application/pdf":
        text = extract_text_from_pdf(file_bytes)
        return None, text, display_info

    # ---- Word ----
    if "wordprocessingml" in mime_type:
        text = extract_text_from_docx(file_bytes)
        return None, text, display_info

    # ---- Excel ----
    if "spreadsheetml" in mime_type:
        text = extract_text_from_xlsx(file_bytes)
        return None, text, display_info

    # ---- 텍스트 기반 파일 ----
    try:
        text = file_bytes.decode("utf-8")
        return None, text, display_info
    except UnicodeDecodeError:
        try:
            text = file_bytes.decode("euc-kr")
            return None, text, display_info
        except Exception:
            return None, f"[파일을 읽을 수 없습니다: {file_name}]", display_info


def format_file_size(size_bytes: int) -> str:
    """파일 크기를 읽기 좋은 형태로 변환"""
    if size_bytes < 1024:
        return f"{size_bytes} B"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    elif size_bytes < 1024 * 1024 * 1024:
        return f"{size_bytes / (1024*1024):.1f} MB"
    else:
        return f"{size_bytes / (1024*1024*1024):.2f} GB"


def build_claude_messages(user_text: str, file_contents: list) -> list:
    """
    사용자 텍스트와 파일 내용을 Claude API 메시지 형태로 조합
    
    file_contents: list of (claude_content_block_or_None, extracted_text, display_info)
    """
    content_blocks = []

    # 이미지 블록 먼저 추가
    for claude_block, extracted_text, _ in file_contents:
        if claude_block is not None:
            content_blocks.append(claude_block)

    # 텍스트 조합
    text_parts = []

    # 문서 텍스트 추가
    for claude_block, extracted_text, display_info in file_contents:
        if claude_block is None and extracted_text:
            text_parts.append(f"[첨부 파일 내용]\n{extracted_text}")

    # 사용자 질문 추가
    if user_text.strip():
        text_parts.append(user_text.strip())

    combined_text = "\n\n".join(text_parts) if text_parts else "첨부된 파일을 분석해주세요."
    content_blocks.append({"type": "text", "text": combined_text})

    return content_blocks
